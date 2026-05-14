"""Dataset preview / download / metadata, served from the same Postgres."""
import csv
import io
import json
import logging
import re
import time
from datetime import date, datetime
from decimal import Decimal
from enum import Enum
from typing import Any, List, Optional

from fastapi import APIRouter, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

from auth import user_email
from database import get_conn

router = APIRouter(prefix="/data", tags=["data-access"])
logger = logging.getLogger(__name__)

DEFAULT_PAGE_SIZE = 10
MAX_PAGE_SIZE = 100
EXCEL_MAX_ROWS = 100_000
PSQL_SCHEMA = "opendata"


class DownloadFormat(str, Enum):
    CSV = "csv"
    JSON = "json"
    XLSX = "xlsx"


class ColumnMetadata(BaseModel):
    name: str
    type: str
    comment: Optional[str] = None


class TableMetadataResponse(BaseModel):
    columns: List[ColumnMetadata]
    table_name: str
    row_count: Optional[int] = None


class PaginationInfo(BaseModel):
    page: int
    page_size: int
    total_count: int
    total_pages: int


class PreviewResponse(BaseModel):
    data: List[dict]
    columns: List[str]
    pagination: PaginationInfo


def _sanitize_identifier(name: str) -> str:
    if not re.match(r"^[a-zA-Z_][a-zA-Z0-9_]*$", name):
        raise ValueError(f"Invalid identifier: {name}")
    return name


def _serialize(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    if isinstance(value, Decimal):
        return float(value)
    return value


def _resolve_table_name(cur, dataset_id: int) -> str:
    cur.execute(
        "SELECT title, psql_src FROM opendata.datasets WHERE id = %s",
        (dataset_id,),
    )
    row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="数据集不存在")
    if not row["psql_src"]:
        raise HTTPException(status_code=404, detail="数据集来源未配置 (psql_src 为空)")
    _sanitize_identifier(row["psql_src"])
    return row["psql_src"], row["title"]


def _fetch_columns(cur, table_name: str) -> List[ColumnMetadata]:
    cur.execute(
        """
        SELECT column_name, data_type,
               col_description(
                   (SELECT oid FROM pg_class
                    WHERE relname = %s
                      AND relnamespace = (SELECT oid FROM pg_namespace WHERE nspname = %s)),
                   ordinal_position
               ) AS column_comment
        FROM information_schema.columns
        WHERE table_schema = %s AND table_name = %s
        ORDER BY ordinal_position
        """,
        (table_name, PSQL_SCHEMA, PSQL_SCHEMA, table_name),
    )
    return [
        ColumnMetadata(name=r["column_name"], type=r["data_type"], comment=r["column_comment"])
        for r in cur.fetchall()
    ]


@router.get("/datasets/{dataset_id}/metadata", response_model=TableMetadataResponse)
def get_metadata(dataset_id: int):
    with get_conn() as conn:
        with conn.cursor() as cur:
            table_name, _ = _resolve_table_name(cur, dataset_id)
            columns = _fetch_columns(cur, table_name)
            if not columns:
                raise HTTPException(
                    status_code=404, detail=f"表 {PSQL_SCHEMA}.{table_name} 不存在或无列信息"
                )
            cur.execute(f'SELECT COUNT(*) AS n FROM "{PSQL_SCHEMA}"."{table_name}"')
            row_count = cur.fetchone()["n"]
    return TableMetadataResponse(
        columns=columns,
        table_name=f"{PSQL_SCHEMA}.{table_name}",
        row_count=row_count,
    )


@router.get("/datasets/{dataset_id}/preview", response_model=PreviewResponse)
def preview(
    dataset_id: int,
    page: int = Query(1, ge=1),
    page_size: int = Query(DEFAULT_PAGE_SIZE, ge=1, le=MAX_PAGE_SIZE),
    columns: Optional[str] = Query(None, description="Comma-separated column subset"),
):
    with get_conn() as conn:
        with conn.cursor() as cur:
            table_name, _ = _resolve_table_name(cur, dataset_id)
            all_cols = [c.name for c in _fetch_columns(cur, table_name)]
            if not all_cols:
                raise HTTPException(
                    status_code=404, detail=f"表 {PSQL_SCHEMA}.{table_name} 不存在"
                )

            if columns:
                col_list = [c.strip() for c in columns.split(",")]
                if any(c not in all_cols for c in col_list):
                    raise HTTPException(status_code=400, detail="包含无效的列名")
                select_cols = ", ".join(f'"{c}"' for c in col_list)
                display_cols = col_list
            else:
                select_cols = "*"
                display_cols = all_cols

            full_name = f'"{PSQL_SCHEMA}"."{table_name}"'
            cur.execute(f"SELECT COUNT(*) AS n FROM {full_name}")
            total = cur.fetchone()["n"]
            total_pages = (total + page_size - 1) // page_size if total else 0
            offset = (page - 1) * page_size

            cur.execute(
                f"SELECT {select_cols} FROM {full_name} LIMIT %s OFFSET %s",
                (page_size, offset),
            )
            rows = cur.fetchall()
            data = [
                {k: _serialize(v) for k, v in row.items()}
                for row in rows
            ]

    return PreviewResponse(
        data=data,
        columns=display_cols,
        pagination=PaginationInfo(
            page=page, page_size=page_size, total_count=total, total_pages=total_pages
        ),
    )


def _csv_iter(rows, column_names):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(column_names)
    yield output.getvalue()
    for row in rows:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([_serialize(row[c]) for c in column_names])
        yield output.getvalue()


def _json_iter(rows, column_names):
    yield "["
    first = True
    for row in rows:
        if not first:
            yield ","
        first = False
        yield json.dumps(
            {c: _serialize(row[c]) for c in column_names},
            ensure_ascii=False,
            default=str,
        )
    yield "]"


def _generate_excel(rows, column_names) -> bytes:
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    for col_idx, col_name in enumerate(column_names, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = cell.font.copy(bold=True)
    for row_idx, row in enumerate(rows, 2):
        for col_idx, col_name in enumerate(column_names, 1):
            ws.cell(row=row_idx, column=col_idx, value=_serialize(row[col_name]))
    for col_idx, col_name in enumerate(column_names, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max(len(str(col_name)) + 2, 10), 50
        )
    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _log_download(
    conn,
    cur,
    user_id: Optional[int],
    dataset_id: int,
    fmt: str,
    file_size: int,
    duration_ms: int,
    ip: Optional[str],
):
    try:
        cur.execute(
            """
            INSERT INTO opendata.download_logs
                (user_id, dataset_id, download_format, file_size_bytes,
                 duration_ms, ip_address)
            VALUES (%s, %s, %s, %s, %s, %s)
            """,
            (user_id, dataset_id, fmt, file_size, duration_ms, ip),
        )
        cur.execute(
            "UPDATE opendata.datasets SET download_count = COALESCE(download_count, 0) + 1 WHERE id = %s",
            (dataset_id,),
        )
        conn.commit()
    except Exception as e:
        logger.warning("Failed to log download: %s", e)
        conn.rollback()


def _resolve_user_id(cur, email: str) -> Optional[int]:
    cur.execute("SELECT id FROM opendata.users WHERE email = %s", (email,))
    row = cur.fetchone()
    return row["id"] if row else None


@router.get("/datasets/{dataset_id}/download")
def download(
    dataset_id: int,
    request: Request,
    format: DownloadFormat = Query(DownloadFormat.CSV),
    columns: Optional[str] = Query(None),
):
    start = time.time()
    email = user_email(request)
    ip = request.client.host if request.client else None

    with get_conn() as conn:
        with conn.cursor() as cur:
            user_id = _resolve_user_id(cur, email)
            table_name, title = _resolve_table_name(cur, dataset_id)
            all_cols = [c.name for c in _fetch_columns(cur, table_name)]
            if not all_cols:
                raise HTTPException(status_code=404, detail="表不存在")

            if columns:
                col_list = [c.strip() for c in columns.split(",")]
                if any(c not in all_cols for c in col_list):
                    raise HTTPException(status_code=400, detail="包含无效的列名")
                select_cols = ", ".join(f'"{c}"' for c in col_list)
            else:
                col_list = all_cols
                select_cols = "*"

            full_name = f'"{PSQL_SCHEMA}"."{table_name}"'
            if format == DownloadFormat.XLSX:
                cur.execute(f"SELECT {select_cols} FROM {full_name} LIMIT {EXCEL_MAX_ROWS}")
            else:
                cur.execute(f"SELECT {select_cols} FROM {full_name}")
            rows = cur.fetchall()
            column_names = col_list if columns else list(rows[0].keys()) if rows else all_cols

            safe_title = re.sub(r"[^\w\-_]", "_", title or f"dataset-{dataset_id}")[:50]
            filename = f"{safe_title}.{format.value}"
            duration_ms = int((time.time() - start) * 1000)

            if format == DownloadFormat.CSV:
                body = "".join(_csv_iter(rows, column_names))
                size = len(body.encode("utf-8"))
                _log_download(conn, cur, user_id, dataset_id, format.value, size, duration_ms, ip)
                return StreamingResponse(
                    iter([body]),
                    media_type="text/csv; charset=utf-8",
                    headers={"Content-Disposition": f'attachment; filename="{filename}"'},
                )
            if format == DownloadFormat.JSON:
                body = "".join(_json_iter(rows, column_names))
                size = len(body.encode("utf-8"))
                _log_download(conn, cur, user_id, dataset_id, format.value, size, duration_ms, ip)
                return StreamingResponse(
                    iter([body]),
                    media_type="application/json; charset=utf-8",
                    headers={"Content-Disposition": f'attachment; filename="{filename}"'},
                )
            # XLSX
            data = _generate_excel(rows, column_names)
            _log_download(conn, cur, user_id, dataset_id, format.value, len(data), duration_ms, ip)
            return StreamingResponse(
                iter([data]),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": f'attachment; filename="{filename}"',
                    "Content-Length": str(len(data)),
                },
            )
